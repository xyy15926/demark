#!/usr/bin/env python3
#----------------------------------------------------------
#   Name: openxl.py
#   Author: xyy15926
#   Created at: 2020-11-17 08:38:33
#   Updated at: 2020-11-17 08:54:09
#   Description: Notes for studying openpyxl, including some
#     code.
#----------------------------------------------------------

# %% [markdown]
# # OpenPyXl

import openpyxl as pyxl
from openpyxl.styles import (Font, Alignment, Border,
        Side, PatternFill, Protection, NamedStyle)
from openpyxl.chart import (Reference, Series, BarChart)
from openpyxl.drawing import (image, )
from openpyxl.styles.colors import (Color, )
from openpyxl.worksheet.table import (Table, TableStyleInfo, )
import os
ROOT_PATH = os.path.abspath("..")

# %% [markdown]
#
# ## Basic Elements
#
# ### Workbook
#
# 1. Get, create workbook
# 2. Get, create, remove worksheet

# Workbook
if os.path.isfile("tmp.xlsx"):
    # load workbook
    wb = pyxl.load_workbook("tmp.xlsx", data_only=True)
else:
    wb = pyxl.Workbook()
# get sheet names
print(wb.get_sheet_names(), wb.sheetnames)

# Creat sheet
ws = wb.create_sheet("Sheet", 1)
ws2 = wb.create_sheet("Sheet2", 1)
ws3 = wb.create_sheet("Sheet3", 1)
# Get sheet with name
ws = wb.get_sheet_by_name("Sheet")
# Get work current active sheet
ws = wb.active
# remove sheet
wb.remove_sheet(ws2)

# %% [markdown]
#
# ### Worksheet
#
# 1. Worksheet dimensions(range)
# 2. Worksheet title(name)

# Set sheet propertie
ws.sheet_properties.tabColor="0072BA"

# Worksheet's some attribute
print(ws.dimensions)
print(ws.title)

# %% [markdown]
#
# ### Cell, Row, Column
#
# 1. Cell: get, set value for cell, cell-slice
# 2. Row: iterate all rows or rows of specified range
# 3. Column:

# Set and get cell's value
ws["A1"] = "haha"
ws.cell(row=1, column=2).value = "purple"
print(ws["A1"].value, ws.cell(row=1, column=2).value)

# Append new row with value provided
data = [
    ['Item', 'Colour'],
    ['pen', 'brown'],
    ['book', 'black'],
    ['plate', 'white'],
    ['chair', 'brown'],
    ['coin', 'gold'],
    ['bed', 'brown'],
    ['notebook', 'white']
]
for row in data:
    ws.append(row)

# Get cell slice
cells = ws["A3": "C3"]
for row in cells:
    for cell in row:
        print(cell.value)

# Iterate row, col
for row in ws.rows:
    for cell in row:
        print(cell.value, end=" ")
for col in ws.columns:
    for cell in col:
        print(cell.value, end=" ")
for row in ws.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
for col in ws.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in col:
        print(cell.value, end=" ")

# Get row, column with label
# Used to set attribute for row or column
print(ws.column_dimensions["A"])
print(ws.row_dimensions[1])

# %% [markdown]
#
# ## Structures and Functions
#
# ### Merge cells
# 
# 1. 合并单元格值、格式仅能在**左上角**上单元格定义

ws.merge_cells("D1:E4")
cell = ws.cell(row=1, column=4)
cell.value = "Sunny"
cell.alignment = pyxl.styles.Alignment(horizontal="center", vertical="center")

# %% [markdown]
#
# ### Filter

ws.auto_filter.ref = "A1:B7"
ws.auto_filter.add_filter_column(1, ["brown", "white"])
ws.auto_filter.add_sort_condition("B2:B8")

# %% [markdown]
#
# ### Freeze

# freeze rows and columns before cell specified below
ws.freeze_panes = "B2"

# %% [markdown]
#
# ### Formula, hyperlink, etc

# Formula
ws["G1"].value = "=SUM(A1:B2)"
ws["G1"].font = ws["A14"].font + \
    pyxl.styles.Font(bold=True)
# Hyperlink with excel formula
ws["G2"].value = "=HYPERLINK('%s', '%s')'" % ("https://baidu.com", "Baidu")
# Hyperlink with openpyxl hyperlink property
ws["G3"].hyperlink = "https://baidu.com"
ws["G3"].value = "Baidu"


# %% [markdown]
#
# ### Image

img = image.Image(os.path.join(ROOT_PATH, "resources/tmp.png"))
ws.add_image(img, "C1")

# %% [markdown]
#
# ### Chart

rows = [
    ("USA", 46),
    ("China", 38),
    ("UK", 29),
    ("Russia", 22),
    ("South Korea", 13),
    ("Germany", 11)
]
for row in rows:
    ws.append(row)
cats = Reference(ws, min_col=1, min_row=10, max_col=1, max_row=16)
data = Reference(ws, min_col=2, min_row=10, max_col=2, max_row=16)
chart = BarChart()
chart.add_data(data=data)
chart.set_categories(labels=cats)
chart.legend = None
chart.y_axis.majorGridlines = None
chart.title = "Gold Medals"
ws.add_chart(chart, "H5")

# %% [markdown]
#
# ## Style
#
# ### Worksheet样式

# hide gridline
ws.sheet_view.showGridLines = False
# set column width
ws.column_dimensions["A"].width = 20
ws.row_dimensions[1].height = 40
# hide columns from "D" to "G"
ws.column_dimensions.group("D", "G", hidden=True)

# %% [markdown]
#
# ### Table样式

table = Table(displayName="Gold", ref="A15:B20")
tb_style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=True,
    showLastColumn=True,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = tb_style
ws.add_table(table)


# %% [markdown] 
#
# ### 单元样式
#
# - 单元样式：即直接创建样式对象
#   - 在对象间共享：一旦被分配，不能再修改，以避免副作用
#   - 应用风格
#     - 活动sheet：需对相应单元格应用样式
#     - 关闭sheet：可以行、列应用样式
# 

font = Font(
    # name="Calibri",
    name="微软雅黑",
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color="FF000000"
)
fill = PatternFill(
    # `patternType` alias
    fill_type=None,
    # `fgColor` alias
    start_color="FFFFFFFF",
    # `bgColor` alias
    end_color="FF000000",
)
border = Border(
    left=Side(
        border_style=None,
        color="FF000000"
    ),
    right=Side(
        border_style=None,
        color="FF000000"
    ),
    top=Side(
        border_style=None,
        color="FF000000"
    ),
    bottom=Side(
        border_style=None,
        color="FF000000"
    ),
    diagonal=Side(
        border_style=None,
        color="FF000000"
    ),
    diagonal_direction=0,
    outline=Side(
        border_style=None,
        color="FF000000"
    ),
    vertical=Side(
        border_style=None,
        color="FF000000"
    ),
    horizontal=Side(
        border_style=None,
        color="FF000000"
    )
)
alignment = Alignment(
    horizontal="general",
    vertical="bottom",
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0
)
number_format="General"
protection=Protection(
    locked=True,
    hidden=False
)
# fill_type 
ws["A1"].fill = PatternFill(
    patternType="solid",
    start_color="00ff00",
    end_color="0000ff"
)

# %% [markdown]
#
# ### 命名样式
#
# - 命名样式
#   - 命名样式可以在workbook中注册：命令样式在首次分配
#     给单元格时自动注册
#   - 注册之后的命名样式可以使用名称指定样式
# - 内置命名样式
#   - 规范中中包含一些内置规范样式，但是样式名称存储在
#     本地化表单中，而Openpyxl只能识别英文名称

hl_style = NamedStyle(name="highlight")
hl_style.font = Font(bold=True, size=20)
bd = Side(style="thick", color="000000")
hl_style.border = Border(
    left=bd,
    top=bd,
    right=bd,
    bottom=bd
)
wb.add_named_style(hl_style)
ws["A1"].style = "highlight"

# %% [markdown]
#
# ### Colors
#
# - `Color`：支持索引颜色、主题、色调

clr = Color(indexed=32)
clr = Color(theme=6, tint=0.5)

# %%
wb.save("tmp.xlsx")
